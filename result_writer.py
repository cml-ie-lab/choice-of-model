import datetime
import os.path
from enum import Enum

import openpyxl as op

import config as cfg
import sut as st


class ResultWriter:

    _env_extension_range = 'B80:BN85'
    _env_extension_label_range = 'A80:A85'

    def __init__(self, fn, sut):
        assert type(fn) is str
        assert os.path.isdir(os.path.dirname(fn))
        assert type(sut) is st.Sut
        self._out_file = fn
        self._sut = sut
        self._cntr_code = sut.cntr_code
        self._yr = sut.year
        self._settings = cfg.Config()
        self._log = None
        self._extensions = None
        self._model_0 = None
        self._model_a = None
        self._model_b = None
        self._model_c = None
        self._model_d = None
        self._model_0_ext = None
        self._model_a_ext = None
        self._model_b_ext = None
        self._model_c_ext = None
        self._model_d_ext = None
        self._wb = None
        self._ws_description = None
        self._ws_log = None
        self._ws_extensions = None
        self._ws_supply = None
        self._ws_dom = None
        self._ws_imp = None
        self._ws_model_0 = None
        self._ws_model_a = None
        self._ws_model_b = None
        self._ws_model_c = None
        self._ws_model_d = None

    def set_env_extensions(self, extensions):
        self._extensions = extensions

    def set_log(self, log):
        self._log = log

    def set_model_0(self, technology_matrix, extensions_matrix):
        self._model_0 = technology_matrix
        self._model_0_ext = extensions_matrix

    def set_model_a(self, technology_matrix, extensions_matrix):
        self._model_a = technology_matrix
        self._model_a_ext = extensions_matrix

    def set_model_b(self, technology_matrix, extensions_matrix):
        self._model_b = technology_matrix
        self._model_b_ext = extensions_matrix

    def set_model_c(self, technology_matrix, extensions_matrix):
        self._model_c = technology_matrix
        self._model_c_ext = extensions_matrix

    def set_model_d(self, technology_matrix, extensions_matrix):
        self._model_d = technology_matrix
        self._model_d_ext = extensions_matrix

    def write_result(self):
        self._wb = op.Workbook()
        self._create_sheets()
        self._write_supply()
        self._write_dom()
        self._write_imp()
        self._write_extensions()
        self._write_model(self._ws_model_0, self._model_0, self._model_0_ext, Format.PXI)
        self._write_model(self._ws_model_a, self._model_a, self._model_a_ext, Format.PXP)
        self._write_model(self._ws_model_b, self._model_b, self._model_b_ext, Format.PXP)
        self._write_model(self._ws_model_c, self._model_c, self._model_c_ext, Format.IXI)
        self._write_model(self._ws_model_d, self._model_d, self._model_d_ext, Format.IXI)
        self._write_log()
        self._write_description()
        self._wb.save(self._out_file)

    def _create_sheets(self):
        self._ws_description = self._wb.create_sheet('description')
        self._ws_supply = self._wb.create_sheet(self._cntr_code + '_sup_' + self._year_str())
        self._ws_dom = self._wb.create_sheet(self._cntr_code + '_dom_' + self._year_str())
        self._ws_imp = self._wb.create_sheet(self._cntr_code + '_imp_' + self._year_str())
        self._ws_extensions = self._wb.create_sheet(self._cntr_code + '_emissions_' + self._year_str())
        self._ws_model_0 = self._wb.create_sheet(self._cntr_code + '_model_0_' + self._year_str())
        self._ws_model_a = self._wb.create_sheet(self._cntr_code + '_model_a_' + self._year_str())
        self._ws_model_b = self._wb.create_sheet(self._cntr_code + '_model_b_' + self._year_str())
        self._ws_model_c = self._wb.create_sheet(self._cntr_code + '_model_c_' + self._year_str())
        self._ws_model_d = self._wb.create_sheet(self._cntr_code + '_model_d_' + self._year_str())
        self._ws_log = self._wb.create_sheet(self._cntr_code + '_log_' + self._year_str())
        del self._wb['Sheet']  # remove default worksheet

    def _write_description(self):
        now = datetime.datetime.now()
        self._ws_description.cell(row=3, column=2, value='Country code:')
        self._ws_description.cell(row=4, column=2, value='Year:')
        self._ws_description.cell(row=5, column=2, value='Created:')
        self._ws_description.cell(row=3, column=3, value=self._cntr_code)
        self._ws_description.cell(row=4, column=3, value=str(self._yr))
        self._ws_description.cell(row=5, column=3, value=now.strftime("%Y-%m-%d %H:%M"))

    def _write_extensions(self):
        sh_range = 'B13:BN18'
        cells = self._ws_extensions[sh_range]
        data = self._extensions
        self._ndarray2cells(data, cells)

        sh_range = self._settings.industry_categories_range
        cells = self._ws_extensions[sh_range]
        data = self._sut.industry_categories
        self._list2cells(data, cells)

        sh_range = self._settings.env_extensions_categories_range
        cells = self._ws_extensions[sh_range]
        data = self._settings.substance_names
        self._list2cells(data, cells)

    def _write_log(self):
        data = self._log
        if len(data) == 0:
            self._ws_log.cell(row=1, column=1, value='no errors logged')
        else:
            for row_idx in range(0, len(data)):
                data_row = data[row_idx]
                self._ws_log.cell(row=row_idx + 1, column=1, value=str(data_row))

    def _write_model(self, sheet, econ_data, env_data, model_format):
        sh_range = self._settings.supply_range
        cells = sheet[sh_range]
        self._ndarray2cells(econ_data, cells)

        sh_range = self._env_extension_range
        cells = sheet[sh_range]
        self._ndarray2cells(env_data, cells)

        sh_range = self._env_extension_label_range
        cells = sheet[sh_range]
        data = self._settings.substance_names
        self._list2cells(data, cells)

        if model_format == Format.PXI:

            sh_range = self._settings.industry_categories_range
            cells = sheet[sh_range]
            econ_data = self._sut.industry_categories
            self._list2cells(econ_data, cells)

            sh_range = self._settings.product_categories_range
            cells = sheet[sh_range]
            econ_data = self._sut.product_categories
            self._list2cells(econ_data, cells)

        if model_format == Format.PXP:

            sh_range = self._settings.industry_categories_range
            cells = sheet[sh_range]
            econ_data = self._sut.product_categories
            self._list2cells(econ_data, cells)

            sh_range = self._settings.product_categories_range
            cells = sheet[sh_range]
            econ_data = self._sut.product_categories
            self._list2cells(econ_data, cells)

        if model_format == Format.IXI:

            sh_range = self._settings.industry_categories_range
            cells = sheet[sh_range]
            econ_data = self._sut.industry_categories
            self._list2cells(econ_data, cells)

            sh_range = self._settings.product_categories_range
            cells = sheet[sh_range]
            econ_data = self._sut.industry_categories
            self._list2cells(econ_data, cells)

    def _write_supply(self):
        sh_range = self._settings.supply_range
        cells = self._ws_supply[sh_range]
        data = self._sut.supply
        self._ndarray2cells(data, cells)

        sh_range = self._settings.industry_categories_range
        cells = self._ws_supply[sh_range]
        data = self._sut.industry_categories
        self._list2cells(data, cells)

        sh_range = self._settings.product_categories_range
        cells = self._ws_supply[sh_range]
        data = self._sut.product_categories
        self._list2cells(data, cells)

    def _write_dom(self):
        sh_range = self._settings.domestic_use_range
        cells = self._ws_dom[sh_range]
        data = self._sut.domestic_use
        self._ndarray2cells(data, cells)

        sh_range = self._settings.domestic_final_use_range
        cells = self._ws_dom[sh_range]
        data = self._sut.domestic_final_use
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.value_added_range
        cells = self._ws_dom[sh_range]
        data = self._sut.value_added
        indices = self._settings.value_added_indices
        self._ndarray2cells(data, cells, row_indices=indices)

        sh_range = self._settings.value_added_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.value_added_categories
        indices = self._settings.value_added_indices
        self._list2cells(data, cells, row_indices=indices)

        sh_range = self._settings.product_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.product_categories
        self._list2cells(data, cells)

        sh_range = self._settings.industry_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.industry_categories
        self._list2cells(data, cells)

        sh_range = self._settings.tax_range
        cells = self._ws_dom[sh_range]
        data = self._sut.tax
        self._ndarray2cells(data, cells)

        sh_range = self._settings.final_tax_range
        cells = self._ws_dom[sh_range]
        data = self._sut.final_tax
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.cif_fob_range
        cells = self._ws_dom[sh_range]
        data = self._sut.cif_fob
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.purchases_non_residents_range
        cells = self._ws_dom[sh_range]
        data = self._sut.purchases_non_residents
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.direct_purchases_abroad_range
        cells = self._ws_dom[sh_range]
        data = self._sut.direct_purchases_abroad
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.tax_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.tax_categories
        self._list2cells(data, cells)

        sh_range = self._settings.cif_fob_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.cif_fob_categories
        self._list2cells(data, cells)

        sh_range = self._settings.purchases_non_residents_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.purchases_non_residents_categories
        self._list2cells(data, cells)

        sh_range = self._settings.direct_purchases_abroad_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.direct_purchases_abroad_categories
        self._list2cells(data, cells)

        sh_range = self._settings.finaluse_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.finaluse_categories
        indices = self._settings.final_use_category_indices
        self._list2cells(data, cells, column_indices=indices)

    def _write_imp(self):
        sh_range = self._settings.import_use_range
        cells = self._ws_imp[sh_range]
        data = self._sut.import_use
        self._ndarray2cells(data, cells)

        sh_range = self._settings.final_import_use_range
        cells = self._ws_imp[sh_range]
        data = self._sut.import_final_use
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.product_categories_range
        cells = self._ws_imp[sh_range]
        data = self._sut.product_categories
        self._list2cells(data, cells)

        sh_range = self._settings.industry_categories_range
        cells = self._ws_imp[sh_range]
        data = self._sut.industry_categories
        self._list2cells(data, cells)

        sh_range = self._settings.finaluse_categories_range
        cells = self._ws_imp[sh_range]
        data = self._sut.finaluse_categories
        indices = self._settings.final_use_category_indices
        self._list2cells(data, cells, column_indices=indices)

    def _year_str(self):
        return str(self._yr)[-2:]

    @staticmethod
    def _ndarray2cells(data, cells, row_indices=None, column_indices=None):
        if row_indices is None and column_indices is None:
            row_idx = 0
            for cell_row in cells:
                col_idx = 0
                for cell in cell_row:
                    cell.value = data[row_idx, col_idx]
                    col_idx += 1
                row_idx += 1

        elif row_indices is not None and column_indices is None:
            cat_idx = 0
            row_idx = 0
            for cell_row in cells:
                col_idx = 0
                if row_idx in row_indices:
                    for cell in cell_row:
                        cell.value = data[cat_idx, col_idx]
                        col_idx += 1
                    cat_idx += 1
                else:
                    for cell in cell_row:
                        cell.value = ':'
                    col_idx += 1
                row_idx += 1

        elif row_indices is None and column_indices is not None:
            row_idx = 0
            for cell_row in cells:
                col_idx = 0
                cat_idx = 0
                for cell in cell_row:
                    if col_idx in column_indices:
                        cell.value = data[row_idx, cat_idx]
                        cat_idx += 1
                    else:
                        cell.value = ':'
                    col_idx += 1
                row_idx += 1
        else:
            raise ValueError('Give either row_indices or column indices, not both')

    @staticmethod
    def _list2cells(data, cells, row_indices=None, column_indices=None):
        if row_indices is None and column_indices is None:
            idx = 0
            for cell_row in cells:
                for cell in cell_row:
                    cell.value = data[idx]
                    idx += 1

        elif row_indices is not None and column_indices is None:
            cat_idx = 0
            row_idx = 0
            for cell_row in cells:
                if row_idx in row_indices:
                    for cell in cell_row:
                        cell.value = data[cat_idx]
                    cat_idx += 1
                else:
                    for cell in cell_row:
                        cell.value = ':'
                row_idx += 1

        elif row_indices is None and column_indices is not None:
            for cell_row in cells:
                col_idx = 0
                cat_idx = 0
                for cell in cell_row:
                    if col_idx in column_indices:
                        cell.value = data[cat_idx]
                        cat_idx += 1
                    else:
                        cell.value = ':'
                    col_idx += 1
        else:
            raise ValueError('Give either row indices or column indices, not both')


class Format(Enum):
    PXI = 1
    PXP = 2
    IXI = 3
