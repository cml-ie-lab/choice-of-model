import numpy as np
import openpyxl as op
import openpyxl.utils.cell as uc

import config as cfg


class EnvReader:

    def __init__(self, code, yr):
        # some checking on the input parameters
        assert type(code) is str
        assert len(code) == 2
        assert type(yr) is int
        self._year = yr
        self._cntrcode = code
        self._settings = cfg.Config()

        # read extensions in native format and concordance file
        env_fn = self._settings.environmental_extensions_file
        concordance_fn = self._settings.concordance_file
        self._env_sh_name = self._cntrcode + '_' + self._year_str()
        self._wb_env = op.load_workbook(env_fn, read_only=True, data_only=True)
        self._wb_concordance = op.load_workbook(concordance_fn, read_only=True, data_only=True)

        concordance = self._get_cordance()
        env_extensions = self._get_env_extensions()

        # match sector classification of extensions file with sector classification supply-use tables
        self._extensions = env_extensions @ concordance

    def get_extensions(self):
        return self._extensions

    def _get_cordance(self):
        sh_range = self._settings.concordance_range
        sh_name = 'concordance'
        ws = self._wb_concordance[sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data

    def _get_env_extensions(self):
        sh_range = self._settings.environmental_extensions_range
        ws = self._wb_env[self._env_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data

    def _cells2ndarray(self, cells, range_string):
        (row_cnt, col_cnt) = self._range_dimensions(range_string)
        data_list = list()
        for cell_row in cells:
            for cell in cell_row:
                if self._is_number(cell.value):
                    data_list.extend([cell.value])
                else:
                    data_list.extend([0])
        data = np.asarray(data_list, dtype=np.float64)
        data = np.reshape(data, (row_cnt, col_cnt))
        return data

    def _cells2list(self, cells, range_string):
        (row_cnt, col_cnt) = self._range_dimensions(range_string)
        if row_cnt == 1 or col_cnt == 1:
            data_list = list()
            for cell_row in cells:
                for cell in cell_row:
                    data_list.extend([cell.value])
        else:
            raise ValueError("cells must be a single row or column")
        return data_list

    def _year_str(self):
        return str(self._year)[-2:]

    @staticmethod
    def _range_dimensions(range_string):
        (min_col, min_row, max_col, max_row) = uc.range_boundaries(range_string)
        row_cnt = max_row - min_row + 1
        col_cnt = max_col - min_col + 1
        return row_cnt, col_cnt

    @staticmethod
    def _is_number(s):
        if s is None:
            return False
        else:
            try:
                float(s)
                return True
            except ValueError:
                return False
