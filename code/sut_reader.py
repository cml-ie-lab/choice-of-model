import os.path

import numpy as np
import openpyxl as op
import openpyxl.utils.cell as uc

import config as cfg
import sut as st


class SutReader:

    def __init__(self, fn, code, yr):
        assert type(code) is str
        assert len(code) == 2
        assert type(yr) is int
        assert type(fn) is str
        assert os.path.isfile(fn)
        self._year = yr
        self._cntrcode = code
        self._fn = fn
        self._settings = cfg.Config()
        self._supply_sh_name = self._cntrcode + '_sup' + self._year_str()
        self._dom_use_sh_name = self._cntrcode + '_dom' + self._year_str()
        self._imp_use_sh_name = self._cntrcode + '_imp' + self._year_str()
        self._log_sh_name = 'log'
        self._wb = op.load_workbook(fn, read_only=True, data_only=True)

    def read_sut(self):
        sut = st.Sut()
        sut.cntr_code = self._cntrcode
        sut.year = self._year
        sut.supply = self._get_supply()
        sut.domestic_use = self._get_domestic_use()
        sut.domestic_final_use = self._get_domestic_final_use()
        sut.import_use = self._get_import_use()
        sut.import_final_use = self._get_final_import_use()
        sut.cif_fob = self._get_cif_fob()
        sut.tax = self._get_tax()
        sut.final_tax = self._get_final_tax()
        sut.direct_purchases_abroad = self._get_direct_purchases_abroad()
        sut.purchases_non_residents = self._get_purchases_non_residents()
        sut.value_added = self._get_value_added()
        sut.product_categories = self._get_product_categories()
        sut.industry_categories = self._get_industry_categories()
        sut.finaluse_categories = self._get_finaluse_categories()
        sut.value_added_categories = self._get_value_added_categories()
        sut.cif_fob_categories = self._get_cif_fob_categories()
        sut.direct_purchases_abroad_categories = self._get_direct_purchases_abroad_categories()
        sut.tax_categories = self._get_tax_categories()
        sut.purchases_non_residents_categories = self._get_direct_purchases_abroad_categories()
        return sut

    def _get_cif_fob_categories(self):
        sh_range = self._settings.cif_fob_categories_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        return data

    def _get_direct_purchases_abroad_categories(self):
        sh_range = self._settings.direct_purchases_abroad_categories_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        return data

    def _get_tax_categories(self):
        sh_range = self._settings.tax_categories_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        return data

    def _get_purchases_non_residents_categories(self):
        sh_range = self._settings.purchases_non_residents_categories_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        return data

    def _get_product_categories(self):
        sh_range = self._settings.product_categories_range
        ws = self._wb[self._supply_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        return data

    def _get_industry_categories(self):
        sh_range = self._settings.industry_categories_range
        ws = self._wb[self._supply_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        return data

    def _get_finaluse_categories(self):
        sh_range = self._settings.finaluse_categories_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        selected_data = [data[index] for index in self._settings.final_use_category_indices]
        return selected_data

    def _get_value_added_categories(self):
        sh_range = self._settings.value_added_categories_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2list(cells, sh_range)
        selected_data = [data[index] for index in self._settings.value_added_indices]
        return selected_data

    def _get_supply(self):
        sh_range = self._settings.supply_range
        ws = self._wb[self._supply_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data

    def _get_domestic_use(self):
        sh_range = self._settings.domestic_use_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data

    def _get_domestic_final_use(self):
        sh_range = self._settings.domestic_final_use_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data[:, self._settings.final_use_category_indices]

    def _get_import_use(self):
        sh_range = self._settings.import_use_range
        ws = self._wb[self._imp_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data

    def _get_final_import_use(self):
        sh_range = self._settings.final_import_use_range
        ws = self._wb[self._imp_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data[:, self._settings.final_use_category_indices]

    def _get_cif_fob(self):
        sh_range = self._settings.cif_fob_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data[:, self._settings.final_use_category_indices]

    def _get_tax(self):
        sh_range = self._settings.tax_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data

    def _get_final_tax(self):
        sh_range = self._settings.final_tax_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data[:, self._settings.final_use_category_indices]

    def _get_direct_purchases_abroad(self):
        sh_range = self._settings.direct_purchases_abroad_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data[:, self._settings.final_use_category_indices]

    def _get_purchases_non_residents(self):
        sh_range = self._settings.purchases_non_residents_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data[:, self._settings.final_use_category_indices]

    def _get_value_added(self):
        sh_range = self._settings.value_added_range
        ws = self._wb[self._dom_use_sh_name]
        cells = ws[sh_range]
        data = self._cells2ndarray(cells, sh_range)
        return data[self._settings.value_added_indices, :]

    def _cells2ndarray(self, cells, range_string):
        (row_cnt, col_cnt) = self._range_dimensions(range_string)
        data_list = list()
        for cell_row in cells:
            for cell in cell_row:
                if self._is_number(cell.value):
                    data_list.extend([cell.value])
                else:
                    data_list.extend([0])
        data = np.asarray(data_list, dtype=np.float64)
        data = np.reshape(data, (row_cnt, col_cnt))
        return data

    def _cells2list(self, cells, range_string):
        (row_cnt, col_cnt) = self._range_dimensions(range_string)
        if row_cnt == 1 or col_cnt == 1:
            data_list = list()
            for cell_row in cells:
                for cell in cell_row:
                    data_list.extend([cell.value])
        else:
            raise ValueError("cells must be a single row or column")
        return data_list

    def _year_str(self):
        return str(self._year)[-2:]

    @staticmethod
    def _range_dimensions(range_string):
        (min_col, min_row, max_col, max_row) = uc.range_boundaries(range_string)
        row_cnt = max_row - min_row + 1
        col_cnt = max_col - min_col + 1
        return row_cnt, col_cnt

    @staticmethod
    def _is_number(s):
        if s is None:
            return False
        else:
            try:
                float(s)
                return True
            except ValueError:
                return False
