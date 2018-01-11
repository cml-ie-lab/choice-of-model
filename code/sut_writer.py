import os.path

import openpyxl as op

import config as cfg
import sut as st


class SutWriter:

    def __init__(self, fn, sut):
        assert type(fn) is str
        assert os.path.isdir(os.path.dirname(fn))
        assert type(sut) is st.Sut
        self._out_file = fn
        self._sut = sut
        self._cntr_code = sut.cntr_code
        self._yr = sut.year
        self._settings = cfg.Config()
        self._wb = None
        self._ws_description = None
        self._ws_supply = None
        self._ws_dom = None
        self._ws_imp = None

    def write_sut(self):
        self._wb = op.Workbook()
        self._create_sheets()
        self._write_supply()
        self._write_dom()
        self._write_imp()
        self._wb.save(self._out_file)

    def _create_sheets(self):
        self._ws_description = self._wb.create_sheet('description')
        self._ws_supply = self._wb.create_sheet(self._cntr_code + '_sup' + self._year_str())
        self._ws_dom = self._wb.create_sheet(self._cntr_code + '_dom' + self._year_str())
        self._ws_imp = self._wb.create_sheet(self._cntr_code + '_imp' + self._year_str())
        del self._wb['Sheet']  # remove default worksheet

    def _write_supply(self):
        sh_range = self._settings.supply_range
        cells = self._ws_supply[sh_range]
        data = self._sut.supply
        self._ndarray2cells(data, cells)

        sh_range = self._settings.industry_categories_range
        cells = self._ws_supply[sh_range]
        data = self._sut.industry_categories
        self._list2cells(data, cells)

        sh_range = self._settings.product_categories_range
        cells = self._ws_supply[sh_range]
        data = self._sut.product_categories
        self._list2cells(data, cells)

    def _write_dom(self):
        sh_range = self._settings.domestic_use_range
        cells = self._ws_dom[sh_range]
        data = self._sut.domestic_use
        self._ndarray2cells(data, cells)

        sh_range = self._settings.domestic_final_use_range
        cells = self._ws_dom[sh_range]
        data = self._sut.domestic_final_use
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.value_added_range
        cells = self._ws_dom[sh_range]
        data = self._sut.value_added
        indices = self._settings.value_added_indices
        self._ndarray2cells(data, cells, row_indices=indices)

        sh_range = self._settings.value_added_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.value_added_categories
        indices = self._settings.value_added_indices
        self._list2cells(data, cells, row_indices=indices)

        sh_range = self._settings.product_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.product_categories
        self._list2cells(data, cells)

        sh_range = self._settings.industry_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.industry_categories
        self._list2cells(data, cells)

        sh_range = self._settings.tax_range
        cells = self._ws_dom[sh_range]
        data = self._sut.tax
        self._ndarray2cells(data, cells)

        sh_range = self._settings.final_tax_range
        cells = self._ws_dom[sh_range]
        data = self._sut.final_tax
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.cif_fob_range
        cells = self._ws_dom[sh_range]
        data = self._sut.cif_fob
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.purchases_non_residents_range
        cells = self._ws_dom[sh_range]
        data = self._sut.purchases_non_residents
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.direct_purchases_abroad_range
        cells = self._ws_dom[sh_range]
        data = self._sut.direct_purchases_abroad
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.tax_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.tax_categories
        self._list2cells(data, cells)

        sh_range = self._settings.cif_fob_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.cif_fob_categories
        self._list2cells(data, cells)

        sh_range = self._settings.purchases_non_residents_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.purchases_non_residents_categories
        self._list2cells(data, cells)

        sh_range = self._settings.direct_purchases_abroad_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.direct_purchases_abroad_categories
        self._list2cells(data, cells)

        sh_range = self._settings.finaluse_categories_range
        cells = self._ws_dom[sh_range]
        data = self._sut.finaluse_categories
        indices = self._settings.final_use_category_indices
        self._list2cells(data, cells, column_indices=indices)

    def _write_imp(self):
        sh_range = self._settings.import_use_range
        cells = self._ws_imp[sh_range]
        data = self._sut.import_use
        self._ndarray2cells(data, cells)

        sh_range = self._settings.final_import_use_range
        cells = self._ws_imp[sh_range]
        data = self._sut.import_final_use
        indices = self._settings.final_use_category_indices
        self._ndarray2cells(data, cells, column_indices=indices)

        sh_range = self._settings.product_categories_range
        cells = self._ws_imp[sh_range]
        data = self._sut.product_categories
        self._list2cells(data, cells)

        sh_range = self._settings.industry_categories_range
        cells = self._ws_imp[sh_range]
        data = self._sut.industry_categories
        self._list2cells(data, cells)

        sh_range = self._settings.finaluse_categories_range
        cells = self._ws_imp[sh_range]
        data = self._sut.finaluse_categories
        indices = self._settings.final_use_category_indices
        self._list2cells(data, cells, column_indices=indices)

    def _year_str(self):
        return str(self._yr)[-2:]

    @staticmethod
    def _ndarray2cells(data, cells, row_indices=None, column_indices=None):
        if row_indices is None and column_indices is None:
            row_idx = 0
            for cell_row in cells:
                col_idx = 0
                for cell in cell_row:
                    cell.value = data[row_idx, col_idx]
                    col_idx += 1
                row_idx += 1

        elif row_indices is not None and column_indices is None:
            cat_idx = 0
            row_idx = 0
            for cell_row in cells:
                col_idx = 0
                if row_idx in row_indices:
                    for cell in cell_row:
                        cell.value = data[cat_idx, col_idx]
                        col_idx += 1
                    cat_idx += 1
                else:
                    for cell in cell_row:
                        cell.value = ':'
                    col_idx += 1
                row_idx += 1

        elif row_indices is None and column_indices is not None:
            row_idx = 0
            for cell_row in cells:
                col_idx = 0
                cat_idx = 0
                for cell in cell_row:
                    if col_idx in column_indices:
                        cell.value = data[row_idx, cat_idx]
                        cat_idx += 1
                    else:
                        cell.value = ':'
                    col_idx += 1
                row_idx += 1
        else:
            raise ValueError('Give either row_indices or column indices, not both')

    @staticmethod
    def _list2cells(data, cells, row_indices=None, column_indices=None):
        if row_indices is None and column_indices is None:
            idx = 0
            for cell_row in cells:
                for cell in cell_row:
                    cell.value = data[idx]
                    idx += 1

        elif row_indices is not None and column_indices is None:
            cat_idx = 0
            row_idx = 0
            for cell_row in cells:
                if row_idx in row_indices:
                    for cell in cell_row:
                        cell.value = data[cat_idx]
                    cat_idx += 1
                else:
                    for cell in cell_row:
                        cell.value = ':'
                row_idx += 1

        elif row_indices is None and column_indices is not None:
            for cell_row in cells:
                col_idx = 0
                cat_idx = 0
                for cell in cell_row:
                    if col_idx in column_indices:
                        cell.value = data[cat_idx]
                        cat_idx += 1
                    else:
                        cell.value = ':'
                    col_idx += 1
        else:
            raise ValueError('Give either row_indices or column indices, not both')
